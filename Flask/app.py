from flask import Flask, request, url_for, render_template, jsonify
from flask_restful import Resource, Api
import requests
import json
import os
import subprocess
from openpyxl import load_workbook
import paypalrestsdk
import logging
from time import sleep

def work_load():
	wb1 = load_workbook('sheet.xlsx')
	ws = wb1["Sheet1"]
	return ws, wb1

def run_motor(Motor):
	import motor1
	motor1.setup(Motor)
	motor1.run(Motor)


ws, wb1 = work_load()

a = ws['B1']   
b = ws['B2']
c = ws['B3']
d = ws['B4']
e = ws['B5']
f = ws['B6']
g = ws['B7']
h = ws['B8']
i = ws['B9']
j = ws['B10']

items = [a,b,c,d,e,f,g,h,i,j]

def restart_trans():
	for i in range (1,11):
		ws.cell(row=i, column=2).value=0
	wb1.save('sheet.xlsx')

def compute_sum():
	work_load()
	c = 0
	for i in range (1,11):
		c = c + ws.cell(row=i, column=2).value * ws.cell(row=i, column=3).value
	return c

restart_trans()

os.environ['http_proxy'] = "http://10.8.0.1:8080" 
os.environ['https_proxy'] = "https://10.8.0.1:8080"

app = Flask(__name__)
api = Api(app)

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def pag_not_found(e):
    return render_template('500.html'), 500

@app.route('/')
def hello():
	restart_trans()
	return render_template('index.html')

@app.route('/sos')
def sos():
   return render_template('sos.html')

@app.route('/purchase')
def purchase():
	return render_template('menu.html')

@app.route('/dis')
def discard():
	restart_trans()
	return render_template('menu.html')

@app.route('/nxt')
def grandt():	
	return render_template('total.html', variable=str(compute_sum()))

@app.route('/med1')
def med1():
	ws.cell(row=1, column=2).value = ws.cell(row=1, column=2).value + 1
	wb1.save('sheet.xlsx')
	return render_template('menu.html')

@app.route('/med2')
def med2():
	ws.cell(row=2, column=2).value = ws.cell(row=2, column=2).value + 1
	wb1.save('sheet.xlsx')
	return render_template('menu.html')

@app.route('/med3')
def med3():
	ws.cell(row=3, column=2).value = ws.cell(row=3, column=2).value + 1
	wb1.save('sheet.xlsx')
	return render_template('menu.html')

@app.route('/med4')
def med4():
	ws.cell(row=4, column=2).value = ws.cell(row=4, column=2).value + 1
	wb1.save('sheet.xlsx')
	return render_template('menu.html')

@app.route('/med5')
def med5():
	ws.cell(row=5, column=2).value = ws.cell(row=5, column=2).value + 1
	wb1.save('sheet.xlsx')
	return render_template('menu.html')

@app.route('/med6')
def med6():
	ws.cell(row=6, column=2).value = ws.cell(row=6, column=2).value + 1
	wb1.save('sheet.xlsx')
	return render_template('menu.html')

@app.route('/med7')
def med7():
	ws.cell(row=7, column=2).value = ws.cell(row=7, column=2).value + 1
	wb1.save('sheet.xlsx')
	return render_template('menu.html')

@app.route('/med8')
def med8():
	ws.cell(row=8, column=2).value = ws.cell(row=8, column=2).value + 1
	wb1.save('sheet.xlsx')
	return render_template('menu.html')

@app.route('/med9')
def med9():
	ws.cell(row=9, column=2).value = ws.cell(row=9, column=2).value + 1
	wb1.save('sheet.xlsx')
	return render_template('menu.html')

@app.route('/med10')
def med10():
	ws.cell(row=10, column=2).value = ws.cell(row=10, column=2).value + 1
	wb1.save('sheet.xlsx')
	return render_template('menu.html')

@app.route('/camera')
def camera():
   return render_template('camera.html')

@app.route('/motor1')
def run_m1():
   import motor1
   motor1.setup(Motor=1)
   motor1.run(Motor=1)
   return render_template('index.html')

@app.route('/motor2')
def run_m2():
   import motor1
   motor1.setup(Motor=2)
   motor1.run(Motor=2)
   return render_template('index.html')

@app.route('/motor3')
def run_m3():
   import motor1
   motor1.setup(Motor=3)
   motor1.run(Motor=3)
   return render_template('index.html')

@app.route('/motor4')
def run_m4():
   import motor1
   motor1.setup(Motor=4)
   motor1.run(Motor=4)
   return render_template('index.html')

@app.route('/motor5')
def run_m5():
   import motor1
   motor1.setup(Motor=5)
   motor1.run(Motor=5)
   return render_template('index.html')

@app.route('/motor6')
def run_m6():
   import motor1
   motor1.setup(Motor=6)
   motor1.run(Motor=6)
   return render_template('index.html')

@app.route('/motor7')
def run_m7():
   import motor1
   motor1.setup(Motor=7)
   motor1.run(Motor=7)
   return render_template('index.html')

@app.route('/motor8')
def run_m8():
   import motor1
   motor1.setup(Motor=8)
   motor1.run(Motor=8)
   return render_template('index.html')

@app.route('/motor9')
def run_m9():
   import motor1
   motor1.setup(Motor=9)
   motor1.run(Motor=9)
   return render_template('index.html')

@app.route('/motor10')
def run_m10():
   import motor1
   motor1.setup(Motor=10)
   motor1.run(Motor=10)
   return render_template('index.html')

@app.route('/afterpay')
def dispense():
	gate = False
	for i in range (1,11):
		for  j in range (1,ws.cell(row=i, column=2).value + 1):
			run_motor(Motor=i)
			sleep(6)
	return render_template('index.html')


@app.route('/test')
def test():
	ws.cell(row=6, column=2).value=5
	wb1.save('sheet.xlsx')
	return

@app.route('/about')
def about():
   return render_template('about.html')

@app.route('/pay')
def pay():
   return render_template('pay.html')

@app.route('/payy')
def payy():
   return '''
    <link rel="stylesheet" href="https://maxcdn.bootstrapcdn.com/bootstrap/3.3.7/css/bootstrap.min.css">
    <script src="https://ajax.googleapis.com/ajax/libs/jquery/3.3.1/jquery.min.js"></script>
    <link href='https://fonts.googleapis.com/css?family=Source+Sans+Pro:400,700,300' rel='stylesheet' type='text/css'>

<div class="jumbotron text-center">
        <h1>Welcome To Payment Gateway</h1>
    <!--
        <button id="paytm" class="smoothScroll btn btn-info btn-lg " >Pay By Paytm</button>
        <iframe name="paytm-frame" frameborder="20" height="500" width="700" hidden></iframe>
        <br></br>
    <script>
        $(document).ready(function(){
            console.log("jq")
            $("#paytm").click(function(){
            $("#paytm-frame").show()      
            $("#paytm-frame").attr('src', 'http://p-y.tm/z8aQm4rRk')
            });
        });
    </script>
    -->

  <script src="https://www.paypalobjects.com/api/checkout.js"></script>
  <div id="paypal-button-container"></div>

  <script>
      paypal.Button.render({

          env: 'sandbox', // sandbox | production

          // Show the buyer a 'Pay Now' button in the checkout flow
          commit: true,

          // payment() is called when the button is clicked
          payment: function() {

              // Set up a url on your server to create the payment
              var CREATE_URL = 'http://localhost:5001/payment';

              // Make a call to your server to set up the payment
              return paypal.request.post(CREATE_URL)
                  .then(function(res) {
                      return res.paymentID;
                  });
          },

          // onAuthorize() is called when the buyer approves the payment
          onAuthorize: function(data, actions) {

              // Set up a url on your server to execute the payment
              var EXECUTE_URL = 'http://localhost:5001/execute';

              // Set up the data you need to pass to your server
              var data = {
                  paymentID: data.paymentID,
                  payerID: data.payerID
              };

              // Make a call to your server to execute the payment
              return paypal.request.post(EXECUTE_URL, data)
                  .then(function (res) {
                     // window.alert('Payment Complete! You will be redirected to homepage now.');
                      window.setTimeout(function(){ window.location.href = "http://localhost:5001"; },3000);
                  });
          }

      }, '#paypal-button-container');
  </script>
</div>

   '''
paypalrestsdk.configure({
  "mode": "sandbox", # sandbox or live
  "client_id": "AR92SyiTyQfPWGs8bH2xAOncMsKiuTXWECm7aBODm62jYBwboMLyyaKmDGBckKT0oDFMpj47AwYOKLuR",
  "client_secret": "EHLV7tym7p9ZU3gr6SrNTwM5BTIWGhYPCZfVyI6J_XfV6kPPjQQ_YTjadR67UwV0DaUMJsIeQGL2x0dh"
  
#   "mode": "live", # sandbox or live
#   "client_id": "Aa1OvenQntqBU_APRyS-4Exd89f69l4_yQwlLHd7hn3Yq1TNoQph_XBgx_1d15YNuMiCPB9FAlYa0_jn",
#   "client_secret": "EPplWCfFfvJVaup1YQ05FRpSdfAPreLZUhrezdgpxRlZ4p8LL3NDd-_xYg-V5C45gLdUredMsCuhnMb0"
  
   })
  
@app.route('/payment', methods=['POST'])
def payment():
    total_amount_to_be_paid = compute_sum()
    print (total_amount_to_be_paid)
    payment = paypalrestsdk.Payment({
    "intent": "sale",
    "payer": {
        "payment_method": "paypal"},
    "redirect_urls": {
        "return_url": "http://localhost:5001",
        "cancel_url": "http://localhost:3000"},
    "transactions": [{
        "item_list": {
            "items": [{
                "name": "Total Payable Amount",
                "sku": "item",
                "price": total_amount_to_be_paid,
                "currency": "INR",
                "quantity": 1}]},
        "amount": {
            "total": total_amount_to_be_paid,
            "currency": "INR"},
        "description": "This is the payment transaction description."}]})

    if payment.create():
        print("Payment created successfully")
    else:
        print(payment.error)
    return jsonify({'paymentID' : payment.id})

@app.route('/execute', methods=['POST'])
def execute():
    success = False
    payment = paypalrestsdk.Payment.find(request.form['paymentID'])

    if payment.execute({'payer_id':request.form['payerID']}):
        print ('Execute Success')
        success = True

    return jsonify({'success' : success})

@app.route('/home')
def home():
    return render_template('home.html')


if __name__ == '__main__':
    app.run(port=5001, debug=True)